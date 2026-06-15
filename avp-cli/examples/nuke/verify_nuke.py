#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Verify the nuke is fair: the exact ground-truth answers must be recoverable
purely from nuke.xlsx. Re-reads the workbook, parses the messy amount strings,
joins entity->segment from the Entities tab, applies the intercompany and D&A
exclusions, and checks all four answers against ground_truth.json.

This is also the reference "correct" computation a Code Mode agent should match.

Run: uv run verify_nuke.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent


def parse_amount(s) -> float:
    """Recover the exact dollar value from a messy cell string."""
    t = str(s).strip().replace(",", "").replace("USD", "").replace("$", "").strip()
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("()")
    val = float(t)
    return -val if neg else val


def main() -> None:
    wb = load_workbook(HERE / "nuke.xlsx", data_only=True, read_only=True)

    # entity -> segment from the Entities tab
    ent = wb["Entities"]
    entity_segment = {}
    for row in ent.iter_rows(min_row=4, values_only=True):
        if row and row[0] and row[1]:
            entity_segment[str(row[0]).strip()] = str(row[1]).strip()

    # P&L detail rows; header is on row 4, data below
    pl = wb["P&L FY2025"]
    rows = list(pl.iter_rows(min_row=5, values_only=True))

    from collections import defaultdict
    net_rev = defaultdict(float)
    cogs = defaultdict(float)
    opex = defaultdict(float)
    q4_opex_all = 0.0
    parsed = 0

    for r in rows:
        if not r or r[1] is None or r[2] is None or r[3] is None:
            continue
        entity, account, amount_s, ic = r[1], r[2], r[3], r[4]
        seg = entity_segment.get(str(entity).strip())
        if seg is None:        # skips TOTAL / GRAND TOTAL / subtotal trap rows
            continue
        amt = parse_amount(amount_s)
        is_ic = str(ic).strip().upper() == "Y"
        parsed += 1
        if account == "Revenue" and not is_ic:
            net_rev[seg] += amt
        elif account == "COGS":
            cogs[seg] += amt
        elif account == "OpEx":
            opex[seg] += amt

    # Q4 OpEx needs the month; re-read with the date column
    for r in rows:
        if not r or r[1] is None or r[2] != "OpEx":
            continue
        if entity_segment.get(str(r[1]).strip()) is None:
            continue
        month = month_of(r[0])
        if month in (10, 11, 12):
            q4_opex_all += parse_amount(r[3])

    ebitda = {s: net_rev[s] - cogs[s] - opex[s] for s in net_rev}
    computed = {
        "maritime-net-revenue": round(net_rev["Maritime"]),
        "aviation-ebitda": round(ebitda["Aviation"]),
        "top-ebitda-segment": max(ebitda, key=ebitda.get),
        "q4-opex-all": round(q4_opex_all),
    }

    gt = json.loads((HERE / "ground_truth.json").read_text())
    expected = {q["id"]: q["expected"]["answer"] for q in gt["questions"]}

    print(f"parsed {parsed} detail rows from P&L\n")
    ok = True
    for qid, exp in expected.items():
        got = computed[qid]
        match = got == exp
        ok = ok and match
        print(f"  {'PASS' if match else 'FAIL'}  {qid}: computed={got} expected={exp}")
    print("\n" + ("ALL PASS — answers are recoverable from the sheet" if ok else "MISMATCH — fix before running the eval"))


def month_of(cell) -> int:
    """Pull the month out of any of the messy date formats."""
    s = str(cell).strip()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return int(m.group(2))
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return int(m.group(1))
    m = re.match(r"\d{1,2}-([A-Za-z]{3})-\d{4}", s)
    months = {mo: i for i, mo in enumerate(
        ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}
    if m:
        return months[m.group(1)]
    for name, idx in months.items():
        if s.startswith(name) or s.startswith(_full(idx)):
            return idx
    return -1


def _full(idx: int) -> str:
    return ["", "January", "February", "March", "April", "May", "June", "July",
            "August", "September", "October", "November", "December"][idx]


if __name__ == "__main__":
    main()
