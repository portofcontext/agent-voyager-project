#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["arcade-mcp-server>=1.17,<2", "openpyxl>=3.1"]
# ///
"""Arcade-style Sheets MCP server over the Context Nuke workbook.

Mirrors the surface of a real spreadsheet tool (the kind Arcade ships): list the
tabs, then read a tab's values. Built on `arcade-mcp-server` (the same framework
Arcade's own tools use), so the demo's tool layer is honest about "Arcade tools."

The trap is structural: `get_values` returns the WHOLE tab by default. A naive
tool-calling agent pulls all 5,700+ rows into its context and then tries to
parse, join, and sum in its head. A Code Mode agent (reaching this same server
through pctx) writes code that calls these tools and computes in a sandbox, so
only the answer comes back.

Workbook path: $NUKE_XLSX, else ./nuke.xlsx next to this file.

    uv run sheets_mcp.py             # stdio (how an agent / pctx spawns it)
    uv run sheets_mcp.py http        # HTTP streaming
    uv run sheets_mcp.py selftest    # no MCP: print tabs + a sample (local check)
"""

import json
import sys
from os import getenv
from pathlib import Path
from typing import Annotated, Literal, cast

from arcade_mcp_server import MCPApp
from openpyxl import load_workbook

WORKBOOK = Path(getenv("NUKE_XLSX", str(Path(__file__).parent / "nuke.xlsx")))

app = MCPApp(name="nuke_sheets", version="0.1.0", log_level="WARNING")


def _wb():
    return load_workbook(WORKBOOK, data_only=True, read_only=True)


@app.tool
def list_sheets() -> Annotated[list[str], "The tab names in the workbook"]:
    """List the worksheet (tab) names in the spreadsheet."""
    return _wb().sheetnames


@app.tool
def get_values(
    sheet_name: Annotated[str, "The tab to read, e.g. 'P&L FY2025'"],
    a1_range: Annotated[str, "Optional A1 range like 'A4:E2000'; empty = the whole tab"] = "",
) -> Annotated[str, "Tab values as tab-separated rows (one row per line)"]:
    """Return the cell values of a tab as TSV text.

    With no range, returns the entire used range of the tab. This is the honest
    surface of a real sheets tool: the data comes back as-is, formatting warts
    and all.
    """
    ws = _wb()[sheet_name]
    cells = ws[a1_range] if a1_range else ws.iter_rows()
    lines = []
    for row in cells:
        lines.append("\t".join("" if c.value is None else str(c.value) for c in row))
    return "\n".join(lines)


def _selftest() -> None:
    wb = _wb()
    print(f"workbook: {WORKBOOK}")
    print(f"tabs: {wb.sheetnames}")
    ws = wb["P&L FY2025"]
    print(f"P&L used range: {ws.calculate_dimension() if not ws.max_row else f'{ws.max_row} rows x {ws.max_column} cols'}")
    print("first data rows:")
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=8, values_only=True)):
        print("  " + " | ".join("" if v is None else str(v) for v in row))


def main() -> None:
    mode = sys.argv[1] if len(sys.argv) > 1 else getenv("AVP_TEST_MCP_TRANSPORT", "stdio")
    if mode == "selftest":
        _selftest()
        return
    app.run(transport=cast(Literal["stdio", "http"], mode))


if __name__ == "__main__":
    main()
