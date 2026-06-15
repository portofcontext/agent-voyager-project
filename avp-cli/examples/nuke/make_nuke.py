#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Generate the Context Nuke: a deliberately brutal FY2025 financial workbook.

The point of the nuke is to make a naive tool-calling agent confidently wrong
while a Code Mode agent (which computes over the data in code) gets the exact
answer every time. We build a CLEAN ledger with known totals, compute ground
truth from it, then render a MESSY workbook that hides those numbers behind:

  - entity -> segment mapping split into a separate tab (forces a join)
  - amounts stored as strings: "$1,234.56", "(2,345.00)" negatives, "1.2M"
  - dates in five different formats
  - a D&A account that must be EXCLUDED from EBITDA
  - intercompany revenue rows that must be ELIMINATED
  - a hardcoded "TOTAL" trap row with the wrong number
  - thousands of rows, so dumping the sheet into context floods the window

Outputs, next to this script:
  - nuke.xlsx           the workbook the agent sees
  - ground_truth.json   the exact answers, computed from the clean ledger

Run: uv run make_nuke.py
"""

from __future__ import annotations

import json
import random
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SEED = 7
HERE = Path(__file__).parent

# ---------------------------------------------------------------------------
# 1. The clean model (this is the source of truth; the xlsx only obscures it)
# ---------------------------------------------------------------------------

# entity -> segment. Segment is NOT shown on the P&L tab; the agent must join.
ENTITIES = {
    "Harbor Freight Co": "Maritime",
    "Tideline Logistics": "Maritime",
    "Blue Channel Shipping": "Maritime",
    "Skyward Cargo": "Aviation",
    "Aileron Air": "Aviation",
    "Ironline Rail": "Rail",
    "Continental Track": "Rail",
    "Helios Power": "Energy",
    "Delta Grid": "Energy",
}

MONTHS = [date(2025, m, 1) for m in range(1, 13)]

# Operating accounts. D&A is operating-adjacent but EXCLUDED from EBITDA.
# COGS and OpEx are stored as positive expense magnitudes.
ACCOUNTS = ["Revenue", "COGS", "OpEx", "D&A"]


def _split_cents(total: int, k: int, rng: random.Random) -> list[int]:
    """Split an integer-cents total into k positive parts that sum exactly."""
    if k <= 1 or total <= 0:
        return [total]
    cuts = sorted(rng.randint(1, total - 1) for _ in range(k - 1))
    parts, prev = [], 0
    for cut in cuts:
        parts.append(cut - prev)
        prev = cut
    parts.append(total - prev)
    return [p for p in parts if p > 0] or [total]


def build_ledger() -> list[dict]:
    """Deterministic clean ledger at transaction grain.

    Monthly figures per entity/account are fixed (so ground truth is stable),
    then split into many individual transactions so the sheet becomes a real
    context bomb (thousands of rows) without changing any total.
    """
    rng = random.Random(SEED)
    rows: list[dict] = []
    # rough per-entity monthly scale so segments separate cleanly on EBITDA
    scale = {
        "Maritime": 900_000,
        "Aviation": 1_400_000,
        "Rail": 600_000,
        "Energy": 1_100_000,
    }
    # transactions per entity-month, per account (the context-bomb multiplier)
    txns = {"Revenue": (15, 30), "COGS": (10, 22), "OpEx": (6, 16), "D&A": (1, 3)}

    def emit(mdate, entity, account, total_cents, intercompany=False):
        lo, hi = txns[account]
        for part in _split_cents(total_cents, rng.randint(lo, hi), rng):
            rows.append(_row(mdate, entity, account, part, intercompany))

    for entity in ENTITIES:
        base = scale[ENTITIES[entity]]
        for mdate in MONTHS:
            rev = int(rng.uniform(0.7, 1.3) * base * 100)          # cents
            emit(mdate, entity, "Revenue", rev)
            emit(mdate, entity, "COGS", int(rev * rng.uniform(0.45, 0.60)))
            emit(mdate, entity, "OpEx", int(rev * rng.uniform(0.12, 0.22)))
            emit(mdate, entity, "D&A", int(rev * rng.uniform(0.04, 0.08)))
            # some months carry intercompany sales that must be eliminated
            if rng.random() < 0.6:
                ic = int(rng.uniform(0.05, 0.15) * base * 100)
                for part in _split_cents(ic, rng.randint(1, 4), rng):
                    rows.append(_row(mdate, entity, "Revenue", part, intercompany=True))
    rng.shuffle(rows)  # interleave so the sheet isn't conveniently grouped
    return rows


def _row(mdate: date, entity: str, account: str, cents: int, intercompany: bool = False) -> dict:
    return {
        "date": mdate,
        "entity": entity,
        "segment": ENTITIES[entity],
        "account": account,
        "cents": cents,
        "intercompany": intercompany,
    }


# ---------------------------------------------------------------------------
# 2. Ground truth (computed from the clean ledger, in cents, then to dollars)
# ---------------------------------------------------------------------------

def compute_ground_truth(ledger: list[dict]) -> dict:
    segs = sorted(set(ENTITIES.values()))

    def total(pred) -> int:
        return sum(r["cents"] for r in ledger if pred(r))

    net_rev = {
        s: total(lambda r, s=s: r["segment"] == s and r["account"] == "Revenue" and not r["intercompany"])
        for s in segs
    }
    cogs = {s: total(lambda r, s=s: r["segment"] == s and r["account"] == "COGS") for s in segs}
    opex = {s: total(lambda r, s=s: r["segment"] == s and r["account"] == "OpEx") for s in segs}
    ebitda = {s: net_rev[s] - cogs[s] - opex[s] for s in segs}

    q4_opex_all = total(lambda r: r["account"] == "OpEx" and r["date"].month in (10, 11, 12))

    def dollars(cents: int) -> int:
        return round(cents / 100)

    top_segment = max(ebitda, key=ebitda.get)

    return {
        "questions": [
            {
                "id": "maritime-net-revenue",
                "prompt": "Using the workbook, compute total FY2025 NET revenue for the Maritime segment: all Revenue for Maritime entities, excluding intercompany sales. Map entities to segments using the Entities tab. Answer in whole US dollars.",
                "expected": {"answer": dollars(net_rev["Maritime"])},
            },
            {
                "id": "aviation-ebitda",
                "prompt": "Using the workbook, compute FY2025 EBITDA for the Aviation segment: net revenue (excluding intercompany) minus COGS minus OpEx. EBITDA EXCLUDES Depreciation & Amortization (D&A). Answer in whole US dollars.",
                "expected": {"answer": dollars(ebitda["Aviation"])},
            },
            {
                "id": "top-ebitda-segment",
                "prompt": "Using the workbook, which segment had the highest FY2025 EBITDA (net revenue excluding intercompany, minus COGS and OpEx, excluding D&A)? Answer with the segment name only.",
                "expected": {"answer": top_segment},
            },
            {
                "id": "q4-opex-all",
                "prompt": "Using the workbook, compute total Q4 2025 (October, November, December) OpEx across ALL segments. Exclude D&A. Answer in whole US dollars.",
                "expected": {"answer": dollars(q4_opex_all)},
            },
        ],
        "_detail": {
            "net_revenue_by_segment": {s: dollars(v) for s, v in net_rev.items()},
            "cogs_by_segment": {s: dollars(v) for s, v in cogs.items()},
            "opex_by_segment": {s: dollars(v) for s, v in opex.items()},
            "ebitda_by_segment": {s: dollars(v) for s, v in ebitda.items()},
            "q4_opex_all_segments": dollars(q4_opex_all),
            "row_count_pnl": len(ledger),
        },
    }


# ---------------------------------------------------------------------------
# 3. The messy rendering (presentation only; never read by ground truth)
# ---------------------------------------------------------------------------

def fmt_amount(cents: int, rng: random.Random) -> str:
    """Render an amount the ugly way but LOSSLESSLY.

    The exact value must always be recoverable from the string (a perfect
    Code Mode agent has to be able to hit the ground truth), so every variant
    keeps full cent precision. The messiness is in the surface form ($ signs,
    commas, spacing, a 'USD' prefix), not in rounding or sign ambiguity.
    """
    d = cents / 100
    return rng.choice([
        f"${d:,.2f}",          # $1,234.56
        f"{d:,.2f}",           # 1,234.56
        f"{d:.2f}",            # 1234.56  (no thousands separators)
        f"${d:,.2f} ",         # trailing space
        f"USD {d:,.2f}",       # currency-code prefix
        f" ${d:,.2f}",         # leading space
    ])


def fmt_date(d: date, rng: random.Random) -> str:
    day = rng.randint(1, 28)
    forms = [
        d.replace(day=day).isoformat(),                         # 2025-01-14
        f"{d.month:02d}/{day:02d}/{d.year}",                    # 01/14/2025
        d.replace(day=day).strftime("%d-%b-%Y"),                # 14-Jan-2025
        d.strftime("%b %Y"),                                    # Jan 2025
        d.replace(day=day).strftime("%B %d, %Y"),               # January 14, 2025
    ]
    return rng.choice(forms)


def write_workbook(ledger: list[dict], gt: dict) -> Path:
    rng = random.Random(SEED + 1)
    wb = Workbook()

    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="012E58")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="184289")
    note_font = Font(italic=True, color="888888")

    # --- P&L tab: the big messy ledger (NO segment column) ---
    ws = wb.active
    ws.title = "P&L FY2025"
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "PortCo Holdings  —  FY2025 Profit & Loss (consolidated detail, UNAUDITED)"
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"] = "Figures in USD. Q4 preliminary. Intercompany not eliminated in this extract."
    ws["A2"].font = note_font

    headers = ["Txn Date", "Entity", "Account", "Amount", "Intercompany?"]
    hr = 4
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=j, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    r = hr + 1
    for i, row in enumerate(ledger):
        ws.cell(row=r, column=1, value=fmt_date(row["date"], rng))
        ws.cell(row=r, column=2, value=row["entity"])
        ws.cell(row=r, column=3, value=row["account"])
        ws.cell(row=r, column=4, value=fmt_amount(row["cents"], rng))
        ws.cell(row=r, column=5, value="Y" if row["intercompany"] else "")
        r += 1
        # drop in a misleading subtotal/total trap every ~600 rows
        if i > 0 and i % 600 == 0:
            ws.cell(row=r, column=2, value="TOTAL (do not use)")
            ws.cell(row=r, column=3, value="Subtotal")
            ws.cell(row=r, column=4, value=f"${rng.uniform(5e6, 9e6):,.2f}").font = Font(bold=True)
            r += 1

    # a final WRONG grand-total trap row
    ws.cell(row=r + 1, column=2, value="GRAND TOTAL")
    ws.cell(row=r + 1, column=3, value="All accounts")
    tot = ws.cell(row=r + 1, column=4, value="$42,000,000.00")
    tot.font = Font(bold=True, color="C00000")
    ws.cell(row=r + 2, column=1, value="* Figures are illustrative and intentionally inconsistent. Do not trust any printed total.").font = note_font

    for j, w in enumerate([16, 22, 12, 16, 14], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # --- Entities tab: the entity -> segment map (the join target) ---
    es = wb.create_sheet("Entities")
    es.merge_cells("A1:C1")
    es["A1"] = "Legal entity directory — map each entity to its reporting segment"
    es["A1"].font = Font(bold=True)
    for j, h in enumerate(["Entity", "Segment", "Status"], start=1):
        cell = es.cell(row=3, column=j, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for i, (entity, segment) in enumerate(sorted(ENTITIES.items()), start=4):
        es.cell(row=i, column=1, value=entity)
        es.cell(row=i, column=2, value=segment)
        es.cell(row=i, column=3, value="Active")
    for j, w in enumerate([24, 14, 10], start=1):
        es.column_dimensions[get_column_letter(j)].width = w

    # --- Notes tab: definitions + a benign red herring ---
    ns = wb.create_sheet("Notes")
    notes = [
        "Reporting notes (FY2025):",
        "1. EBITDA = Net Revenue - COGS - OpEx. EBITDA EXCLUDES Depreciation & Amortization (D&A).",
        "2. Net Revenue excludes intercompany sales (rows flagged Intercompany? = Y).",
        "3. COGS and OpEx are recorded as positive expense magnitudes.",
        "4. Amounts on the P&L tab are formatted inconsistently by the source system.",
        "5. The printed TOTAL / GRAND TOTAL rows are placeholders and are not reliable.",
        "6. Segment is not on the P&L tab; join via the Entities tab.",
    ]
    for i, line in enumerate(notes, start=1):
        ns.cell(row=i, column=1, value=line)
    ns.column_dimensions["A"].width = 100

    out = HERE / "nuke.xlsx"
    wb.save(out)
    return out


def main() -> None:
    ledger = build_ledger()
    gt = compute_ground_truth(ledger)
    xlsx = write_workbook(ledger, gt)
    (HERE / "ground_truth.json").write_text(json.dumps(gt, indent=2))
    print(f"wrote {xlsx} ({xlsx.stat().st_size // 1024} KB)")
    print(f"wrote {HERE / 'ground_truth.json'}")
    print("\nground truth:")
    print(json.dumps(gt["_detail"], indent=2))
    print("\nquestions + expected:")
    for q in gt["questions"]:
        print(f"  [{q['id']}] expected={q['expected']}")


if __name__ == "__main__":
    main()
